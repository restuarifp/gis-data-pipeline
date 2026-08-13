#!/usr/bin/env python3
"""
split_excel.py

Mengambil file .xlsx dari folder sumber di Nextcloud (via WebDAV),
memisahkan setiap sheet menjadi file tersendiri, lalu mengupload
hasilnya ke folder tujuan agar bisa dibaca Airbyte per-sheet.

Naming convention output:
  source: /Uploads/kantorA/Finance/finance.xlsx  (sheet: Pendapatan)
  dest  : kantorA__finance__Pendapatan.xlsx

  nama_kantor diambil dari komponen path tepat di atas subfolder 'Finance'
  (lihat _kantor_name); untuk struktur tanpa 'Finance', dipakai komponen terakhir.

Mode:
  python split_excel.py           -- jalankan sekali lalu keluar
  python split_excel.py A2/Finance [B1/Finance ...]
                                  -- jalankan sekali untuk folder tertentu saja
  python split_excel.py --watch   -- loop berkala (SCHEDULE_INTERVAL_MINUTES)
  python split_excel.py --serve   -- server kontrol HTTP (job_control) di gisnet,
                                     dipakai bot Telegram untuk memicu run

Path sumber:
  NEXTCLOUD_SOURCE_PATHS relatif terhadap NEXTCLOUD_SOURCE_HOME (bila di-set),
  sehingga prefix panjang tidak perlu diulang di tiap entri. Lihat resolve_source().
"""

import io
import logging
import os
import re
import sys
import time
import posixpath
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import PurePosixPath

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from urllib.parse import urlparse, unquote

load_dotenv()

# ── Logging ────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)

# ── Config ─────────────────────────────────────────────────────────────────

NEXTCLOUD_URL      = os.environ["NEXTCLOUD_URL"]           # https://cloud.example.com
NEXTCLOUD_USER     = os.environ["NEXTCLOUD_USER"]
NEXTCLOUD_PASSWORD = os.environ["NEXTCLOUD_PASSWORD"]
DEST_PATH          = unquote(os.environ["NEXTCLOUD_DEST_PATH"].strip())
SCHEDULE_MINUTES   = int(os.getenv("SCHEDULE_INTERVAL_MINUTES", "60"))

# Retry saat file tujuan terkunci (HTTP 423 Locked) — mis. OnlyOffice/klien lain
# sedang memegang lock atas file di folder tujuan.
WEBDAV_MAX_RETRIES           = int(os.getenv("WEBDAV_MAX_RETRIES", "5"))
WEBDAV_RETRY_BACKOFF_SECONDS = float(os.getenv("WEBDAV_RETRY_BACKOFF_SECONDS", "3"))

# Folder induk untuk semua path sumber. Kalau di-set, entri di
# NEXTCLOUD_SOURCE_PATHS cukup ditulis relatif (mis. "A1/Finance").
SOURCE_HOME = unquote(os.getenv("NEXTCLOUD_SOURCE_HOME", "").strip()).strip("/")


def resolve_source(path: str) -> str:
    """
    Ubah satu entri path sumber menjadi path absolut di Nextcloud.

    - SOURCE_HOME kosong  -> path dipakai apa adanya (perilaku sebelum var ini ada).
    - path sudah di bawah SOURCE_HOME -> dibiarkan. Join-nya idempoten, jadi .env
      lama yang menulis prefix lengkap di tiap entri tetap jalan tanpa diedit.
    - selain itu -> digabung di bawah SOURCE_HOME.

    Hasil yang keluar dari SOURCE_HOME (mis. lewat '..') ditolak: path bisa datang
    dari argumen perintah /split di Telegram, dan itu tidak boleh bisa menunjuk ke
    folder sembarang di akun Nextcloud.
    """
    bersih = posixpath.normpath(unquote(path.strip()).replace("\\", "/")).strip("/")
    if bersih in ("", "."):
        raise ValueError("path sumber kosong")

    # normpath tidak membuang '..' yang berada di depan; tolak eksplisit. Cek ini
    # berlaku juga saat SOURCE_HOME kosong, di mana tidak ada batas lain.
    if bersih == ".." or bersih.startswith("../") or "/../" in bersih:
        raise ValueError(f"path {path!r} mengandung '..'")

    if not SOURCE_HOME:
        return f"/{bersih}"

    if bersih == SOURCE_HOME or bersih.startswith(f"{SOURCE_HOME}/"):
        hasil = bersih
    else:
        hasil = posixpath.normpath(f"{SOURCE_HOME}/{bersih}").strip("/")

    if hasil != SOURCE_HOME and not hasil.startswith(f"{SOURCE_HOME}/"):
        raise ValueError(
            f"path {path!r} keluar dari NEXTCLOUD_SOURCE_HOME ({SOURCE_HOME!r})"
        )
    return f"/{hasil}"


def parse_sources(raw: str) -> list[str]:
    """Pecah string koma/newline menjadi daftar path absolut."""
    return [resolve_source(p) for p in re.split(r"[,\n]", raw) if p.strip()]


# Dukung kedua nama variabel (NEXTCLOUD_SOURCE_PATHS dan NEXTCLOUD_SOURCE_PATH).
# Pisahkan dengan koma atau newline; unquote() handle path URL-encoded (%20, dll.).
_raw_sources = (
    os.environ.get("NEXTCLOUD_SOURCE_PATHS")
    or os.environ.get("NEXTCLOUD_SOURCE_PATH")
    or ""
)
if not _raw_sources:
    raise ValueError("Set NEXTCLOUD_SOURCE_PATHS (atau NEXTCLOUD_SOURCE_PATH) di .env")

print(f"[split-excel] RAW SOURCE: {repr(_raw_sources)}", flush=True)
if SOURCE_HOME:
    print(f"[split-excel] SOURCE HOME: /{SOURCE_HOME}", flush=True)
else:
    # Tanpa SOURCE_HOME, argumen /split dari Telegram bisa menunjuk folder mana pun
    # di akun Nextcloud (selain '..', yang tetap ditolak).
    print(
        "[split-excel] PERINGATAN: NEXTCLOUD_SOURCE_HOME kosong — "
        "argumen path pada POST /run tidak dibatasi ke satu folder induk.",
        flush=True,
    )

SOURCE_PATHS: list[str] = parse_sources(_raw_sources)

print(f"[split-excel] PARSED {len(SOURCE_PATHS)} path(s):", flush=True)
for i, p in enumerate(SOURCE_PATHS, 1):
    print(f"[split-excel]   [{i}] {repr(p)}", flush=True)

# Hanya ambil origin (scheme + host) dari NEXTCLOUD_URL,
# menghindari dobel path jika user menyertakan /remote.php/... di URL.
_parsed = urlparse(NEXTCLOUD_URL)
_origin = f"{_parsed.scheme}://{_parsed.netloc}"
WEBDAV_BASE = f"{_origin}/remote.php/dav/files/{NEXTCLOUD_USER}"

SESSION = requests.Session()
SESSION.auth = (NEXTCLOUD_USER, NEXTCLOUD_PASSWORD)
SESSION.headers.update({"Content-Type": "application/xml; charset=utf-8"})

# ── WebDAV helpers ─────────────────────────────────────────────────────────

def _url(remote_path: str) -> str:
    """
    Bangun URL WebDAV lengkap.
    Toleran terhadap path yang diawali /{username}/ (Nextcloud kadang
    menampilkan path dalam format tersebut di UI) — prefix username
    dibuang karena WEBDAV_BASE sudah menyertakannya.
    """
    normalized = remote_path.lstrip("/")
    user_prefix = f"{NEXTCLOUD_USER}/"
    if normalized.startswith(user_prefix):
        normalized = normalized[len(user_prefix):]
    return f"{WEBDAV_BASE}/{normalized}"


def list_xlsx(folder: str) -> list[str]:
    """Kembalikan daftar nama file .xlsx di folder WebDAV (tidak rekursif)."""
    resp = SESSION.request(
        "PROPFIND",
        _url(folder),
        headers={"Depth": "1"},
        data=b"""<?xml version="1.0" encoding="utf-8"?>
<d:propfind xmlns:d="DAV:">
  <d:prop><d:displayname/><d:resourcetype/></d:prop>
</d:propfind>""",
    )
    resp.raise_for_status()

    ns = {"d": "DAV:"}
    tree = ET.fromstring(resp.content)

    # Gunakan _url() agar normalisasi username-prefix konsisten
    folder_href = urlparse(_url(folder)).path.rstrip("/")

    results = []
    for node in tree.findall(".//d:response", ns):
        href = (node.findtext("d:href", namespaces=ns) or "").rstrip("/")
        # href sudah URL-encoded (mis. %20); decode agar cocok dengan nama asli
        name = unquote(href.split("/")[-1])
        # Lewati folder itu sendiri
        if href.endswith(folder_href):
            continue
        # Lewati direktori (resourcetype berisi <d:collection/>) — hanya file yang diproses
        if node.find(".//d:resourcetype/d:collection", ns) is not None:
            continue
        # HANYA file .xlsx asli — abaikan format lain (.xls, .csv, .pdf, dst.)
        if not name.lower().endswith(".xlsx"):
            log.info("    ⏭ Lewati (bukan .xlsx): %s", name)
            continue
        # Abaikan file lock/temp Office (mis. ~$laporan.xlsx) yang bukan workbook valid
        if name.startswith("~$") or name.startswith("."):
            log.info("    ⏭ Lewati (file temp/lock): %s", name)
            continue
        results.append(name)
    return results


def download(folder: str, filename: str) -> bytes:
    path = str(PurePosixPath(folder) / filename)
    resp = SESSION.get(_url(path))
    resp.raise_for_status()
    return resp.content


def ensure_folder(folder: str) -> None:
    """Buat folder via MKCOL; abaikan jika sudah ada (405)."""
    resp = SESSION.request("MKCOL", _url(folder))
    if resp.status_code not in (201, 301, 405):
        resp.raise_for_status()


def _request_with_retry(method: str, url: str, **kwargs) -> requests.Response:
    """
    Jalankan request WebDAV, ulangi khusus saat 423 Locked dengan backoff linear.
    Lock biasanya transient (OnlyOffice melepas setelah beberapa detik), jadi
    retry singkat cukup untuk menembus jendela terkunci tanpa menyisakan file basi.
    """
    resp = None
    for attempt in range(1, WEBDAV_MAX_RETRIES + 1):
        resp = SESSION.request(method, url, **kwargs)
        if resp.status_code != 423:
            return resp
        if attempt < WEBDAV_MAX_RETRIES:
            wait = WEBDAV_RETRY_BACKOFF_SECONDS * attempt
            log.warning(
                "    ⏳ 423 Locked (%s), percobaan %d/%d — tunggu %.0fs: %s",
                method, attempt, WEBDAV_MAX_RETRIES, wait, url,
            )
            time.sleep(wait)
    return resp


def upload(folder: str, filename: str, content: bytes) -> None:
    """Upload/overwrite file via PUT (menimpa file lama tanpa perlu DELETE dulu)."""
    path = str(PurePosixPath(folder) / filename)
    resp = _request_with_retry("PUT", _url(path), data=content)
    resp.raise_for_status()


def delete_file(folder: str, filename: str) -> None:
    path = str(PurePosixPath(folder) / filename)
    resp = _request_with_retry("DELETE", _url(path))
    if resp.status_code not in (200, 204, 404):
        resp.raise_for_status()


def list_xlsx_by_prefix(folder: str, prefix: str) -> list[str]:
    """Kembalikan nama file .xlsx di folder yang diawali prefix tertentu."""
    return [f for f in list_xlsx(folder) if f.startswith(prefix)]


# ── Excel split ─────────────────────────────────────────────────────────────

def _copy_sheet(src, dst) -> list[str]:
    """
    Salin nilai sel, style dasar, lebar kolom, tinggi baris, dan merged cells.
    Kembalikan daftar koordinat sel yang masih berupa formula (bukan nilai) —
    artinya cached value-nya tidak tersedia (file di-save tanpa menyimpan hasil
    hitung formula), sehingga tidak bisa disalin sebagai angka.
    """
    leftover_formulas: list[str] = []
    for row in src.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                leftover_formulas.append(cell.coordinate)
            dst_cell = dst.cell(row=cell.row, column=cell.column, value=value)
            if cell.has_style:
                dst_cell.font           = copy(cell.font)
                dst_cell.border         = copy(cell.border)
                dst_cell.fill           = copy(cell.fill)
                dst_cell.number_format  = cell.number_format
                dst_cell.protection     = copy(cell.protection)
                dst_cell.alignment      = copy(cell.alignment)

    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width

    for row_num, dim in src.row_dimensions.items():
        dst.row_dimensions[row_num].height = dim.height

    for merge in src.merged_cells.ranges:
        dst.merge_cells(str(merge))

    return leftover_formulas


def _safe_name(sheet_name: str) -> str:
    """Ganti karakter tidak aman untuk nama file dengan underscore."""
    return re.sub(r"[^\w\-]", "_", sheet_name).strip("_") or "Sheet"


def split_workbook(content: bytes) -> dict[str, bytes]:
    """
    Pisahkan workbook per sheet.
    Kembalikan dict: {safe_sheet_name: xlsx_bytes}
    data_only=True agar formula diganti nilai tersimpan terakhir (cached value).
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    output: dict[str, bytes] = {}

    for sheet_name in wb.sheetnames:
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name
        leftover_formulas = _copy_sheet(wb[sheet_name], new_ws)
        if leftover_formulas:
            log.warning(
                "    ⚠ Sheet '%s': %d sel formula tanpa cached value (tersimpan "
                "sebagai teks formula, bukan nilai): %s",
                sheet_name, len(leftover_formulas), leftover_formulas[:10],
            )

        buf = io.BytesIO()
        new_wb.save(buf)
        output[_safe_name(sheet_name)] = buf.getvalue()

    return output


# ── Proses utama ────────────────────────────────────────────────────────────

def _kantor_name(source_path: str) -> str:
    """
    Ambil nama kantor dari path folder sumber.

    Struktur folder: <nama kantor>/Finance/finance.xlsx — folder yang di-scan
    adalah <nama kantor>/Finance, sehingga nama kantor = komponen tepat DI ATAS
    subfolder 'Finance'. Jika komponen terakhir bukan 'Finance' (mis. struktur
    capil <nama kantor>/), komponen terakhir dipakai apa adanya.
    """
    parts = [p for p in source_path.strip("/").split("/") if p]
    if not parts:
        return "unknown"
    if len(parts) >= 2 and parts[-1].lower() == "finance":
        return parts[-2]
    return parts[-1]


def process_source(source_path: str) -> bool:
    """
    Proses semua .xlsx di satu folder sumber.
    Kembalikan True jika semua sheet berhasil di-upload, False jika ada kegagalan.
    """
    nama_kantor = _kantor_name(source_path)
    log.info("── Sumber: %s  (kantor: %s)", source_path, nama_kantor)

    try:
        files = list_xlsx(source_path)
    except Exception as exc:
        log.error("  Gagal membaca folder: %s", exc)
        return False

    if not files:
        log.info("  Tidak ada file .xlsx")
        return True

    log.info("  Ditemukan %d file: %s", len(files), files)

    # Kumpulkan semua hasil split ke memory dulu.
    # Jika ada file yang gagal diproses, batalkan seluruh kantor ini
    # daripada menghapus file lama lalu upload sebagian.
    pending: dict[str, bytes] = {}  # {dest_filename: bytes}
    for filename in files:
        stem = filename[:-5]  # hapus .xlsx
        log.info("  Memproses: %s", filename)
        try:
            content = download(source_path, filename)
            sheets = split_workbook(content)
            log.info("    %d sheet: %s", len(sheets), list(sheets))
            for safe_sheet, sheet_bytes in sheets.items():
                pending[f"{nama_kantor}__{stem}__{safe_sheet}.xlsx"] = sheet_bytes
        except Exception as exc:
            log.error("    ✗ Gagal memproses %s: %s", filename, exc, exc_info=True)
            log.error("  Batalkan upload untuk kantor %s", nama_kantor)
            return False

    # Upload (overwrite) DULU sebelum menghapus apa pun. PUT menimpa file lama,
    # jadi kalau ada file terkunci (423) dan gagal walau sudah retry, file lama
    # tetap utuh — kita tidak menyisakan folder kosong sebagian atau data campuran.
    failed: list[str] = []
    for dest_name, sheet_bytes in pending.items():
        try:
            upload(DEST_PATH, dest_name, sheet_bytes)
            log.info("    ✓ Upload: %s", dest_name)
        except Exception as exc:
            failed.append(dest_name)
            log.error("    ✗ Gagal upload %s: %s", dest_name, exc, exc_info=True)

    if failed:
        log.error(
            "  ✗ %d/%d file GAGAL di-upload untuk kantor %s (kemungkinan terkunci "
            "OnlyOffice). File lama TIDAK dihapus agar tidak menyisakan data basi "
            "yang campur baru+lama: %s",
            len(failed), len(pending), nama_kantor, failed,
        )
        return False

    # Semua upload sukses — hapus file lama dengan prefix sama yang TIDAK lagi
    # diproduksi (mis. sheet dihapus/rename di sumber). File yang di-overwrite
    # tidak perlu dihapus karena PUT sudah menimpanya.
    prefix = f"{nama_kantor}__"
    stale = [f for f in list_xlsx_by_prefix(DEST_PATH, prefix) if f not in pending]
    if stale:
        log.info("  Hapus %d file usang: %s", len(stale), stale)
        for old_name in stale:
            try:
                delete_file(DEST_PATH, old_name)
                log.info("    🗑 Dihapus: %s", old_name)
            except Exception as exc:
                log.error("    ✗ Gagal hapus %s: %s", old_name, exc, exc_info=True)

    return True


def run_once(sources: list[str] | None = None) -> bool:
    """
    Jalankan satu siklus. True jika semua sumber sukses.

    `sources` None = pakai SOURCE_PATHS dari .env (perilaku terjadwal). Daftar
    eksplisit dipakai saat run dipicu manual, mis. `/split A1/Finance` dari
    Telegram — path-nya sudah lewat resolve_source() di validate_params().
    """
    targets = sources or SOURCE_PATHS

    log.info("=== Mulai run split-excel ===")
    log.info("  Tujuan  : %s", DEST_PATH)
    log.info("  Sumber  : %d folder%s", len(targets), " (dipilih manual)" if sources else "")
    for i, p in enumerate(targets, 1):
        log.info("    [%d] %s", i, p)

    try:
        ensure_folder(DEST_PATH)
    except Exception as exc:
        log.error("Gagal memastikan folder tujuan ada: %s", exc)
        return False

    ok = True
    for source_path in targets:
        if not process_source(source_path):
            ok = False

    if ok:
        log.info("=== Run selesai (semua sukses) ===")
    else:
        log.error("=== Run selesai DENGAN KEGAGALAN — lihat error di atas ===")
    return ok


def validate_params(params: dict) -> dict:
    """
    Validasi body POST /run. Melempar ValueError -> dibalas 400 oleh job_control.

    Ini satu-satunya tempat path dari luar diperiksa; bot Telegram sengaja tidak
    ikut memvalidasi supaya tidak ada dua aturan yang bisa berbeda.
    """
    raw = params.get("sources")
    if raw in (None, "", []):
        return {}
    if isinstance(raw, str):
        raw = re.split(r"[,\s]+", raw)
    if not isinstance(raw, list):
        raise ValueError("field 'sources' harus berupa list atau string")

    sources = [resolve_source(str(p)) for p in raw if str(p).strip()]
    if not sources:
        raise ValueError("field 'sources' kosong setelah dibersihkan")
    return {"sources": sources}


def main() -> None:
    flag = {a for a in sys.argv[1:] if a.startswith("-")}
    tak_dikenal = flag - {"--watch", "--serve"}
    if tak_dikenal:
        log.error("Opsi tidak dikenal: %s (yang ada: --watch, --serve)",
                  ", ".join(sorted(tak_dikenal)))
        sys.exit(2)

    # Argumen non-flag = daftar folder sumber, seperti argumen /split di Telegram.
    # Dulu argumen begini diabaikan diam-diam sehingga run tetap memakai seluruh
    # NEXTCLOUD_SOURCE_PATHS — terlihat seolah perintahnya tidak berpengaruh.
    argumen = [a for a in sys.argv[1:] if not a.startswith("-")]
    pilihan = None
    if argumen:
        if "--watch" in flag:
            log.error("Argumen folder tidak bisa digabung dengan --watch.")
            sys.exit(2)
        try:
            pilihan = validate_params({"sources": argumen}).get("sources")
        except ValueError as exc:
            log.error("Argumen ditolak: %s", exc)
            sys.exit(2)

    if "--serve" in sys.argv:
        import job_control

        job_control.serve(
            "split-excel",
            lambda params: run_once(params.get("sources")),
            validate=validate_params,
            background="--watch" in sys.argv,
            info={
                "source_home": f"/{SOURCE_HOME}" if SOURCE_HOME else None,
                "sources": SOURCE_PATHS,
                "dest": DEST_PATH,
                "interval_menit": SCHEDULE_MINUTES,
                "fitur": "sources-arg",  # penanda kode ini sudah mendukung /split <path>
            },
        )
        if "--watch" not in sys.argv:
            return

    if "--watch" in sys.argv:
        log.info("Mode watch aktif, interval: %d menit", SCHEDULE_MINUTES)
        while True:
            run_once()
            log.info("Tunggu %d menit...", SCHEDULE_MINUTES)
            time.sleep(SCHEDULE_MINUTES * 60)
    else:
        sys.exit(0 if run_once(pilihan) else 1)


if __name__ == "__main__":
    main()
